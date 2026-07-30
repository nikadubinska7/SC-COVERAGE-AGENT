from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("reports")

DARK = "111827"
DARK_2 = "1F2937"
MID = "374151"
LIGHT = "F3F4F6"
LIGHT_2 = "E5E7EB"
WHITE = "FFFFFF"
GREEN = "16A34A"
AMBER = "F59E0B"
RED = "DC2626"
BLUE = "2563EB"
TEAL = "0F766E"
BORDER = "D1D5DB"


def fmt_number(value: float | int | None) -> str:
    if value is None:
        return "0"
    return f"{float(value):,.0f}"


def fmt_pct(value: float | int | None) -> str:
    if value is None:
        return "0.0%"
    return f"{float(value):.1%}"


def humanize_column_name(name: str) -> str:
    replacements = {
        "eta": "ETA",
        "crd": "CRD",
        "id": "ID",
    }

    label = name.replace("_", " ").title()

    for old, new in replacements.items():
        label = label.replace(old.title(), new)

    label = label.replace("Usd", "USD")
    label = label.replace("Msrp", "MSRP")

    return label


def humanize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    out.columns = [humanize_column_name(str(col)) for col in out.columns]
    return out


def build_summary_dataframe(summary: dict[str, Any], observations: list[str]) -> pd.DataFrame:
    rows = [
        ("Source rows", summary.get("source_rows")),
        ("Included rows", summary.get("included_rows")),
        ("Cancelled rows", summary.get("cancelled_rows")),
        ("Total value", summary.get("total_value")),
        ("Booked/Shipped value", summary.get("booked_shipped_value")),
        ("Available value", summary.get("available_value")),
        ("Open order value", summary.get("open_order_value")),
        ("Covered value", summary.get("covered_value")),
        ("Value coverage %", summary.get("value_coverage_percentage")),
        ("Total volume", summary.get("total_volume")),
        ("Booked/Shipped volume", summary.get("booked_shipped_volume")),
        ("Available volume", summary.get("available_volume")),
        ("Open order volume", summary.get("open_order_volume")),
        ("Covered volume", summary.get("covered_volume")),
        ("Volume coverage %", summary.get("volume_coverage_percentage")),
        ("Risk level", summary.get("risk_level")),
        ("Seasons", ", ".join(summary.get("seasons", []))),
        ("Requested months", ", ".join(summary.get("requested_months", []))),
    ]

    df = pd.DataFrame(rows, columns=["Metric", "Value"])

    observation_rows = pd.DataFrame(
        [
            (f"Observation {idx}", observation)
            for idx, observation in enumerate(observations, start=1)
        ],
        columns=["Metric", "Value"],
    )

    return pd.concat([df, observation_rows], ignore_index=True)


def build_validation_dataframe(validation: dict[str, Any]) -> pd.DataFrame:
    return pd.DataFrame(
        [(key, str(value)) for key, value in validation.items()],
        columns=["Check", "Result"],
    )


def build_filter_options_dataframe(filter_options: dict[str, list[str]]) -> pd.DataFrame:
    if not filter_options:
        return pd.DataFrame()

    max_len = max((len(values) for values in filter_options.values()), default=0)

    data = {}
    for field, values in filter_options.items():
        padded = values + [""] * (max_len - len(values))
        data[humanize_column_name(field)] = padded

    return pd.DataFrame(data)


def build_metric_view_dataframe(coverage_summary: pd.DataFrame, metric: str) -> pd.DataFrame:
    if coverage_summary.empty:
        return pd.DataFrame()

    df = coverage_summary.copy()
    metric_label = metric.title()

    output = pd.DataFrame(
        {
            "Season": df["season"],
            "Requested Month": df["requested_month"],
            "Period": df["season"].astype(str) + " | " + df["requested_month"].astype(str),
            "Booked/Shipped": df[f"booked_shipped_{metric}"],
            "Available": df[f"available_{metric}"],
            "Open Order": df[f"open_order_{metric}"],
            "Covered": df[f"covered_{metric}"],
            f"Total {metric_label}": df[f"total_{metric}"],
            "Coverage %": df[f"{metric}_coverage_percentage"],
            "Open Order %": df[f"open_order_{metric}_percentage"],
        }
    )

    return output.sort_values(["Season", "Requested Month"]).reset_index(drop=True)


def thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def add_excel_table(ws, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 2:
        return

    ws.auto_filter.ref = ws.dimensions


def add_metric_view_chart(ws, title: str) -> None:
    # Keep the workbook Excel-compatible by avoiding drawing/chart XML.
    # The dashboard uses in-cell visual bars instead.
    return


def format_metric_view_sheet(ws, metric: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
            cell.fill = PatternFill("solid", fgColor=LIGHT if cell.row % 2 == 0 else WHITE)

    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value).lower()

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if not isinstance(cell.value, (int, float)):
                continue

            if "%" in header:
                cell.number_format = "0.0%"
            elif metric == "value" and col_idx >= 4:
                cell.number_format = '#,##0'
            elif metric == "volume" and col_idx >= 4:
                cell.number_format = '#,##0'

    for col_idx in range(1, 11):
        ws.column_dimensions[get_column_letter(col_idx)].width = 17

    ws.column_dimensions["C"].width = 22
    add_excel_table(ws, f"{metric}_view")
    add_metric_view_chart(ws, f"{metric.title()} Coverage Mix by Season")


def merge_and_style(
    ws,
    cell_range: str,
    value: Any,
    font: Font,
    fill: PatternFill,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    ws.merge_cells(cell_range)

    start_cell = cell_range.split(":")[0]
    cell = ws[start_cell]
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment or Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    if border:
        for row in ws[cell_range]:
            for merged_cell in row:
                merged_cell.border = border


def section_header(ws, row: int, title: str, start_col: int = 1, end_col: int = 14) -> None:
    cell_range = f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
    merge_and_style(
        ws,
        cell_range,
        title,
        font=Font(color=WHITE, bold=True, size=12),
        fill=PatternFill("solid", fgColor=DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[row].height = 22


def kpi_card(
    ws,
    cell_range: str,
    label: str,
    value: str,
    subtitle: str,
    fill_color: str,
) -> None:
    merge_and_style(
        ws,
        cell_range,
        f"{label}\n{value}\n{subtitle}",
        font=Font(color=WHITE, bold=True, size=13),
        fill=PatternFill("solid", fgColor=fill_color),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin_border(),
    )

    start_row = int("".join(filter(str.isdigit, cell_range.split(":")[0])))
    end_row = int("".join(filter(str.isdigit, cell_range.split(":")[1])))

    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = 24


def visual_bar_row(
    ws,
    row: int,
    label: str,
    value_text: str,
    percentage: float,
    color: str,
    start_col: int = 1,
) -> None:
    label_cell = ws.cell(row=row, column=start_col)
    label_cell.value = label
    label_cell.font = Font(bold=True, color=DARK)
    label_cell.alignment = Alignment(vertical="center")
    label_cell.fill = PatternFill("solid", fgColor=WHITE)
    label_cell.border = thin_border()

    bar_start = start_col + 1
    bar_end = start_col + 8

    safe_pct = max(0, min(float(percentage or 0), 1))
    filled_cols = round((bar_end - bar_start + 1) * safe_pct)

    for col in range(bar_start, bar_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = ""
        cell.border = thin_border()

        if col < bar_start + filled_cols:
            cell.fill = PatternFill("solid", fgColor=color)
        else:
            cell.fill = PatternFill("solid", fgColor=LIGHT_2)

    value_cell = ws.cell(row=row, column=bar_end + 1)
    value_cell.value = value_text
    value_cell.font = Font(bold=True, color=DARK)
    value_cell.alignment = Alignment(horizontal="right", vertical="center")
    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.border = thin_border()

    pct_cell = ws.cell(row=row, column=bar_end + 2)
    pct_cell.value = fmt_pct(safe_pct)
    pct_cell.font = Font(bold=True, color=DARK)
    pct_cell.alignment = Alignment(horizontal="right", vertical="center")
    pct_cell.fill = PatternFill("solid", fgColor=WHITE)
    pct_cell.border = thin_border()

    ws.row_dimensions[row].height = 21


def metric_matrix(
    ws,
    start_row: int,
    title: str,
    rows: list[tuple[str, str, str]],
    start_col: int,
    end_col: int,
) -> None:
    section_header(ws, start_row, title, start_col=start_col, end_col=end_col)

    headers = ["Metric", "Value", "Volume"]
    for idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row + 1, column=idx)
        cell.value = header
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=MID)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    for row_idx, row_data in enumerate(rows, start=start_row + 2):
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=start_col + col_offset)
            cell.value = value
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
            cell.fill = PatternFill("solid", fgColor=WHITE)

            if col_offset == 0:
                cell.font = Font(bold=True, color=DARK)
            else:
                cell.font = Font(color=DARK)

        ws.row_dimensions[row_idx].height = 21


def apply_table_style(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor=DARK)
    header_font = Font(color=WHITE, bold=True)

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            else:
                cell.fill = PatternFill("solid", fgColor=WHITE)

    ws.freeze_panes = f"A{header_row + 1}"

    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions


def autosize_columns(ws, min_width: int = 10, max_width: int = 35) -> None:
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, min_width),
            max_width,
        )


def apply_number_formats(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value).lower()

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if not isinstance(cell.value, (int, float)):
                continue

            if "percentage" in header or "%" in header:
                cell.number_format = "0.0%"
            elif "volume" in header or "quantity" in header or "rows" in header:
                cell.number_format = "#,##0"
            elif "value" in header or "wholesale" in header or "price" in header:
                cell.number_format = "#,##0"
            else:
                cell.number_format = "#,##0"


def format_dashboard(ws) -> None:
    ws.sheet_view.showGridLines = False

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["N"].width = 16

    for row in range(1, 70):
        ws.row_dimensions[row].height = 20


def create_dashboard(ws, report_data: dict[str, Any], observations: list[str]) -> None:
    summary = report_data["executive_summary"]
    coverage_summary = pd.DataFrame(report_data["coverage_summary"])
    timing_risk = pd.DataFrame(report_data["timing_risk"])

    format_dashboard(ws)

    merge_and_style(
        ws,
        "A1:N2",
        "SC COVERAGE REPORT — SNIPES / NIKE-JORDAN",
        font=Font(size=20, bold=True, color=WHITE),
        fill=PatternFill("solid", fgColor=DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
    )

    merge_and_style(
        ws,
        "A4:N5",
        (
            f"Filters applied: Banner = Snipes | "
            f"Seasons = {', '.join(summary.get('seasons', []))} | "
            f"Requested months = {', '.join(summary.get('requested_months', []))} | "
            f"Metric view = Value + Volume"
        ),
        font=Font(size=10, bold=True, color=DARK),
        fill=PatternFill("solid", fgColor=LIGHT),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )

    section_header(ws, 7, "Executive KPIs", start_col=1, end_col=14)

    risk_color = RED
    if summary.get("risk_level") == "Medium":
        risk_color = AMBER
    elif summary.get("risk_level") == "Low":
        risk_color = GREEN

    kpi_card(ws, "A8:C11", "Total Value", fmt_number(summary.get("total_value")), "Wholesale value", DARK_2)
    kpi_card(ws, "D8:F11", "Covered Value", fmt_number(summary.get("covered_value")), f"{fmt_pct(summary.get('value_coverage_percentage'))} coverage", GREEN)
    kpi_card(ws, "G8:I11", "Open Order Value", fmt_number(summary.get("open_order_value")), "Uncovered exposure", RED)
    kpi_card(ws, "J8:L11", "Risk Level", str(summary.get("risk_level", "N/A")), "Based on value coverage", risk_color)

    kpi_card(ws, "A13:C16", "Total Volume", fmt_number(summary.get("total_volume")), "Units", DARK_2)
    kpi_card(ws, "D13:F16", "Covered Volume", fmt_number(summary.get("covered_volume")), f"{fmt_pct(summary.get('volume_coverage_percentage'))} coverage", GREEN)
    kpi_card(ws, "G13:I16", "Open Order Volume", fmt_number(summary.get("open_order_volume")), "Uncovered units", RED)
    kpi_card(ws, "J13:L16", "Included Rows", fmt_number(summary.get("included_rows")), "Report records", BLUE)

    # Left-side sections must end at I to avoid overlapping right-side merged blocks.
    section_header(ws, 19, "Coverage Mix — Value", start_col=1, end_col=9)

    total_value = summary.get("total_value", 0) or 0
    total_volume = summary.get("total_volume", 0) or 0

    mix_value_rows = [
        ("Booked/Shipped", fmt_number(summary.get("booked_shipped_value")), summary.get("booked_shipped_value", 0) / total_value if total_value else 0, GREEN),
        ("Available", fmt_number(summary.get("available_value")), summary.get("available_value", 0) / total_value if total_value else 0, BLUE),
        ("Open Order", fmt_number(summary.get("open_order_value")), summary.get("open_order_value", 0) / total_value if total_value else 0, RED),
    ]

    for idx, (label, value, pct, color) in enumerate(mix_value_rows, start=20):
        visual_bar_row(ws, idx, label, value, pct, color, start_col=1)

    section_header(ws, 25, "Coverage Mix — Volume", start_col=1, end_col=9)

    mix_volume_rows = [
        ("Booked/Shipped", fmt_number(summary.get("booked_shipped_volume")), summary.get("booked_shipped_volume", 0) / total_volume if total_volume else 0, GREEN),
        ("Available", fmt_number(summary.get("available_volume")), summary.get("available_volume", 0) / total_volume if total_volume else 0, BLUE),
        ("Open Order", fmt_number(summary.get("open_order_volume")), summary.get("open_order_volume", 0) / total_volume if total_volume else 0, RED),
    ]

    for idx, (label, value, pct, color) in enumerate(mix_volume_rows, start=26):
        visual_bar_row(ws, idx, label, value, pct, color, start_col=1)

    section_header(ws, 31, "Open Order Timing Risk — Value", start_col=1, end_col=9)

    timing_tot_value = float(
        timing_risk.get("total_open_order_value", pd.Series(dtype=float)).sum()
    )
    timing_tot_volume = float(
        timing_risk.get("total_open_order_volume", pd.Series(dtype=float)).sum()
    )

    timing_rows = [
        ("Early/On Time", float(timing_risk.get("early_on_time_value", pd.Series(dtype=float)).sum()), float(timing_risk.get("early_on_time_volume", pd.Series(dtype=float)).sum()), GREEN),
        ("+1 week", float(timing_risk.get("plus_1_week_value", pd.Series(dtype=float)).sum()), float(timing_risk.get("plus_1_week_volume", pd.Series(dtype=float)).sum()), AMBER),
        ("+2 weeks", float(timing_risk.get("plus_2_weeks_value", pd.Series(dtype=float)).sum()), float(timing_risk.get("plus_2_weeks_volume", pd.Series(dtype=float)).sum()), AMBER),
        ("+3 weeks", float(timing_risk.get("plus_3_weeks_value", pd.Series(dtype=float)).sum()), float(timing_risk.get("plus_3_weeks_volume", pd.Series(dtype=float)).sum()), RED),
        ("+4 weeks or later", float(timing_risk.get("plus_4_weeks_or_later_value", pd.Series(dtype=float)).sum()), float(timing_risk.get("plus_4_weeks_or_later_volume", pd.Series(dtype=float)).sum()), RED),
    ]

    for idx, (label, value, _volume, color) in enumerate(timing_rows, start=32):
        pct = value / timing_tot_value if timing_tot_value else 0
        visual_bar_row(ws, idx, label, fmt_number(value), pct, color, start_col=1)

    section_header(ws, 39, "Open Order Timing Risk — Volume", start_col=1, end_col=9)

    for idx, (label, _value, volume, color) in enumerate(timing_rows, start=40):
        pct = volume / timing_tot_volume if timing_tot_volume else 0
        visual_bar_row(ws, idx, label, fmt_number(volume), pct, color, start_col=1)

    season_rows: list[tuple[str, str, str]] = []

    if not coverage_summary.empty:
        season_data = (
            coverage_summary.groupby("season", as_index=False)
            .agg(
                total_value=("total_value", "sum"),
                total_volume=("total_volume", "sum"),
            )
            .sort_values("season")
        )

        for row in season_data.itertuples(index=False):
            season_rows.append(
                (
                    row.season,
                    fmt_number(row.total_value),
                    fmt_number(row.total_volume),
                )
            )

    metric_matrix(
        ws,
        start_row=19,
        title="Season Overview",
        rows=season_rows,
        start_col=10,
        end_col=14,
    )

    section_header(ws, 31, "Agent Observations", start_col=10, end_col=14)

    for idx, observation in enumerate(observations[:7], start=32):
        merge_and_style(
            ws,
            f"J{idx}:N{idx}",
            f"• {observation}",
            font=Font(size=10, color=DARK),
            fill=PatternFill("solid", fgColor=WHITE),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
            border=thin_border(),
        )
        ws.row_dimensions[idx].height = 36

    validation = report_data["validation"]
    badge_fill = GREEN if validation.get("passes_reconciliation") else RED
    badge_text = "VALIDATION: PASS" if validation.get("passes_reconciliation") else "VALIDATION: FAIL"

    merge_and_style(
        ws,
        "J43:N45",
        (
            f"{badge_text}\n"
            f"Value diff: {validation.get('value_difference')}\n"
            f"Volume diff: {validation.get('volume_difference')}"
        ),
        font=Font(size=12, bold=True, color=WHITE),
        fill=PatternFill("solid", fgColor=badge_fill),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin_border(),
    )


def export_report_to_excel(
    report_data: dict[str, Any],
    observations: list[str],
    output_path: str | Path | None = None,
) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        output_path = OUTPUT_DIR / "sc_coverage_report.xlsx"
    else:
        output_path = Path(output_path)

    summary_df = build_summary_dataframe(
        report_data["executive_summary"],
        observations,
    )

    coverage_summary_df = humanize_dataframe(pd.DataFrame(report_data["coverage_summary"]))
    coverage_summary_raw_df = pd.DataFrame(report_data["coverage_summary"])
    value_view_df = build_metric_view_dataframe(coverage_summary_raw_df, "value")
    volume_view_df = build_metric_view_dataframe(coverage_summary_raw_df, "volume")
    timing_risk_df = humanize_dataframe(pd.DataFrame(report_data["timing_risk"]))
    raw_coverage_df = humanize_dataframe(pd.DataFrame(report_data["coverage_by_season"]))
    filter_options_df = build_filter_options_dataframe(report_data.get("filter_options", {}))
    validation_df = build_validation_dataframe(report_data["validation"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(
            writer,
            sheet_name="Dashboard",
            index=False,
            header=False,
        )

        value_view_df.to_excel(writer, sheet_name="Value View", index=False)
        volume_view_df.to_excel(writer, sheet_name="Volume View", index=False)
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        coverage_summary_df.to_excel(writer, sheet_name="Coverage Summary", index=False)
        timing_risk_df.to_excel(writer, sheet_name="Timing Risk", index=False)
        raw_coverage_df.to_excel(writer, sheet_name="Raw Coverage Data", index=False)
        filter_options_df.to_excel(writer, sheet_name="Filter Options", index=False)
        validation_df.to_excel(writer, sheet_name="Validation", index=False)

        dashboard = writer.sheets["Dashboard"]
        create_dashboard(dashboard, report_data, observations)

        for sheet_name, worksheet in writer.sheets.items():
            if sheet_name in {"Dashboard", "Value View", "Volume View"}:
                continue

            worksheet.sheet_view.showGridLines = False

            if worksheet.max_row > 1:
                apply_table_style(worksheet)
                apply_number_formats(worksheet)
                autosize_columns(worksheet)
                add_excel_table(worksheet, sheet_name)

        format_metric_view_sheet(writer.sheets["Value View"], "value")
        format_metric_view_sheet(writer.sheets["Volume View"], "volume")

        writer.sheets["Executive Summary"].column_dimensions["A"].width = 30
        writer.sheets["Executive Summary"].column_dimensions["B"].width = 85
        writer.sheets["Validation"].column_dimensions["A"].width = 35
        writer.sheets["Validation"].column_dimensions["B"].width = 90

    return output_path
