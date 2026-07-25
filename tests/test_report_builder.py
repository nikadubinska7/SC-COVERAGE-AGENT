from openpyxl import load_workbook

from src.services.report_builder import export_report_to_excel
from src.services.transformations import build_coverage_report


def test_export_report_creates_filterable_metric_views(tmp_path):
    records = [
        {
            "source_row_number": 2,
            "banner": "Snipes",
            "season": "HO2026",
            "status": "Booked/Shipped",
            "order_type": "Standard Order - Futures",
            "requested_month": "202610",
            "confirmed_wholesale": 1000,
            "available_wholesale": 0,
            "report_wholesale_value": 1000,
            "report_volume": 100,
            "coverage_performance": "",
            "eta_vs_crd": "",
        },
        {
            "source_row_number": 3,
            "banner": "Snipes",
            "season": "SP2027",
            "status": "Available",
            "order_type": "Standard Order - Futures",
            "requested_month": "202701",
            "confirmed_wholesale": 0,
            "available_wholesale": 500,
            "report_wholesale_value": 500,
            "report_volume": 50,
            "coverage_performance": "",
            "eta_vs_crd": "",
        },
        {
            "source_row_number": 4,
            "banner": "Snipes",
            "season": "SP2027",
            "status": "Open Order",
            "order_type": "Standard Order - Futures",
            "requested_month": "202701",
            "confirmed_wholesale": 1500,
            "available_wholesale": 0,
            "report_wholesale_value": 1500,
            "report_volume": 150,
            "coverage_performance": "",
            "eta_vs_crd": "10",
        },
    ]

    report = build_coverage_report(records)
    output_path = tmp_path / "report.xlsx"

    export_report_to_excel(
        report_data=report,
        observations=["Filterable report generated."],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert "Value View" in workbook.sheetnames
    assert "Volume View" in workbook.sheetnames
    assert "Dashboard" in workbook.sheetnames

    value_view = workbook["Value View"]
    volume_view = workbook["Volume View"]

    assert value_view.auto_filter.ref == "A1:J3"
    assert volume_view.auto_filter.ref == "A1:J3"
    assert value_view["A1"].value == "Season"
    assert volume_view["A1"].value == "Season"
    assert value_view["H1"].value == "Total Value"
    assert volume_view["H1"].value == "Total Volume"
    assert list(value_view.tables.keys()) == []
    assert list(volume_view.tables.keys()) == []
    assert len(value_view._charts) == 0
    assert len(volume_view._charts) == 0
